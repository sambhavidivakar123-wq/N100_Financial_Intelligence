from pathlib import Path

import pandas as pd

from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule


class PeerComparison:
    def __init__(self):
        self.metrics = [
            "roe",
            "roce",
            "net_profit_margin",
            "free_cash_flow",
            "asset_turnover",
        ]

    def calculate_percentiles(self, df: pd.DataFrame):

        result = df.copy()

        for metric in self.metrics:
            if metric in result.columns:
                result[f"{metric}_percentile"] = (
                    result.groupby("sector")[metric]
                    .rank(method="average", pct=True)
                    * 100
                ).round(2)

        if "debt_to_equity" in result.columns:
            result["debt_to_equity_percentile"] = (
                (
                    1
                    - result.groupby("sector")["debt_to_equity"]
                    .rank(method="average", pct=True)
                )
                * 100
            ).round(2)

        return result

    def export_excel(
        self,
        df: pd.DataFrame,
        output_file="output/peer_comparison.xlsx",
    ):

        Path("output").mkdir(exist_ok=True)

        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:

            for sector in sorted(df["sector"].unique()):

                sector_df = (
                    df[df["sector"] == sector]
                    .sort_values(
                        "roe_percentile",
                        ascending=False,
                    )
                )

                sector_df.to_excel(
                    writer,
                    sheet_name=sector[:31],
                    index=False,
                )

        # Add conditional formatting
        workbook = load_workbook(output_file)

        percentile_rule = ColorScaleRule(
            start_type="min",
            start_color="F8696B",
            mid_type="percentile",
            mid_value=50,
            mid_color="FFEB84",
            end_type="max",
            end_color="63BE7B",
        )

        for worksheet in workbook.worksheets:

            for column in worksheet.iter_cols():

                header = column[0].value

                if header and "percentile" in str(header):

                    column_letter = column[0].column_letter

                    worksheet.conditional_formatting.add(
                        f"{column_letter}2:{column_letter}{worksheet.max_row}",
                        percentile_rule,
                    )

        workbook.save(output_file)

        print(f"Saved: {output_file}")